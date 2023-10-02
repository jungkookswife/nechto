# Как запустить

# py flowers.py

# Установка библиотек

# pip install -r requirements.txt

from bouqet import Flower, Bouquet
from openpyxl import load_workbook

def get_table():
    """
    Прочитать данные из таблицы и вернуть список значений.
    """
    
    wb = load_workbook("./FlowersTable.xlsx")
    sheet_list = wb.get_sheet_names()
    ws = wb[sheet_list[0]] # Открываем "Лист 1"
    
    table = [] # делаем список

    for num in range(1, ws.max_row):
        flower_type = ws[num][1]
        flower_color = ws[num][2]
        flower_length = ws[num][3]
        flower_cost = int(ws[num][4])
        flower_item = int(ws[num][5])

        table.append({
            "type": flower_type, # тип
            "color": flower_color, # цвет
            "length": flower_length, # длина
            "cost": flower_cost, # стоимость
            "item": flower_item # кол-во
        })

    return table # возвращаем список


def main():
    bouquet = Bouquet()
    
    # Получаем информацию от пользователя о количестве и виде цветов
    num_flowers = int(input("Введите количество цветов в букете: "))
    
    for i in range(num_flowers):
        name = input(f"Введите название цветка {i+1}: ")
        color = input(f"Введите цвет цветка {i+1}: ")
        length = int(input(f"Введите длину цветка {i+1}: "))
        
        flower = Flower(name, color, length)
        bouquet.add_flower(flower)
    
    cost = bouquet.calculate_cost()
    print(f"Стоимость букета: {cost} рублей")


if __name__ == "__main__":
    # __name__ == "__main__" = True
    main()
